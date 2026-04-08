import base64
import io
import logging
import re
from odoo import _, api, fields, models
from odoo.exceptions import UserError
from .account_move_line import _CFG_DEFAULTS

_logger = logging.getLogger(__name__)


class HrPayslip(models.Model):
    _inherit = 'hr.payslip'

    # ── CBC Paymaster columns B–T computed from payslip data ─────────────
    # B–D: employee's bank account details
    # E: employee_id.name (use directly in view)
    # F–I, K, L–O, S, T: from jinasena.paymaster.config
    # J: net_wage (use directly in view)
    # P: employee badge / number
    # Q: payslip reference
    # R: date_to as YYMMDD

    x_dest_bank_micr = fields.Char('Dest Bank (B)',    compute='_compute_cbc_payslip', store=False)
    x_dest_branch    = fields.Char('Dest Branch (C)',  compute='_compute_cbc_payslip', store=False)
    x_dest_account   = fields.Char('Dest Account (D)', compute='_compute_cbc_payslip', store=False)
    x_trn_code       = fields.Char('TRN Code (F)',     compute='_compute_cbc_payslip', store=False)
    x_return_code    = fields.Char('Return Code (G)',  compute='_compute_cbc_payslip', store=False)
    x_cr_dr_code     = fields.Char('Cr/Dr (H)',        compute='_compute_cbc_payslip', store=False)
    x_return_date    = fields.Char('Return Date (I)',  compute='_compute_cbc_payslip', store=False)
    x_currency_code  = fields.Char('Currency (K)',     compute='_compute_cbc_payslip', store=False)
    x_orig_bank_micr = fields.Char('Orig Bank (L)',    compute='_compute_cbc_payslip', store=False)
    x_orig_branch    = fields.Char('Orig Branch (M)',  compute='_compute_cbc_payslip', store=False)
    x_orig_account   = fields.Char('Orig Account (N)', compute='_compute_cbc_payslip', store=False)
    x_orig_name      = fields.Char('Orig Name (O)',    compute='_compute_cbc_payslip', store=False)
    x_particulars    = fields.Char('Particulars (P)',  compute='_compute_cbc_payslip', store=False)
    x_cbc_reference  = fields.Char('Reference (Q)',   compute='_compute_cbc_payslip', store=False)
    x_value_date     = fields.Char('Value Date (R)',   compute='_compute_cbc_payslip', store=False)
    x_security_field = fields.Char('Security (S)',     compute='_compute_cbc_payslip', store=False)
    x_filler         = fields.Char('Filler (T)',       compute='_compute_cbc_payslip', store=False)

    @api.depends(
        'employee_id', 'employee_id.bank_account_id',
        'employee_id.bank_account_id.bank_id',
        'net_wage', 'date_to', 'name',
    )
    def _compute_cbc_payslip(self):
        # Load CBC config with fallback defaults
        cfg = _CFG_DEFAULTS.copy()
        try:
            rec = self.env['jinasena.paymaster.config'].sudo().get_config()
            cfg.update({
                'trn_code':       rec.trn_code      or cfg['trn_code'],
                'return_code':    rec.return_code   or cfg['return_code'],
                'cr_dr_code':     rec.cr_dr_code    or cfg['cr_dr_code'],
                'return_date':    rec.return_date   or cfg['return_date'],
                'currency_code':  rec.currency_code or cfg['currency_code'],
                'orig_bank_micr': rec.orig_bank_micr or '',
                'orig_branch':    rec.orig_branch   or '',
                'orig_account':   rec.orig_account  or '',
                'orig_name':      rec.orig_name     or '',
                'security_field': rec.security_field or cfg['security_field'],
                'filler':         rec.filler        or cfg['filler'],
            })
        except Exception:
            _logger.warning(
                'bank_data: jinasena.paymaster.config not available. Using defaults.',
                exc_info=False,
            )

        for slip in self:
            bank = slip.employee_id.bank_account_id

            slip.x_dest_bank_micr = (bank.bank_id.x_micr_code or '') if (bank and bank.bank_id) else ''
            slip.x_dest_branch    = (bank.x_branch_code or '') if bank else ''
            slip.x_dest_account   = (bank.acc_number or '')[:12] if bank else ''

            slip.x_trn_code       = cfg['trn_code']
            slip.x_return_code    = cfg['return_code']
            slip.x_cr_dr_code     = cfg['cr_dr_code']
            slip.x_return_date    = cfg['return_date']
            slip.x_currency_code  = cfg['currency_code']
            slip.x_orig_bank_micr = cfg['orig_bank_micr']
            slip.x_orig_branch    = cfg['orig_branch']
            slip.x_orig_account   = cfg['orig_account']
            slip.x_orig_name      = cfg['orig_name']

            slip.x_particulars    = (slip.employee_id.barcode or str(slip.employee_id.id or ''))[:15]
            slip.x_cbc_reference  = (slip.name or '')[:15]
            slip.x_value_date     = slip.date_to.strftime('%y%m%d') if slip.date_to else ''
            slip.x_security_field = cfg['security_field']
            slip.x_filler         = cfg['filler']

    # ── CBC Paymaster XLSX export — mimics com_upload sheet format ───────

    def action_export_bank_data_csv(self):
        """Export payslips as XLSX matching the CBC Paymaster com_upload sheet."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError(_('openpyxl is required for XLSX export. Please contact your administrator.'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'com_upload'

        # ── Styles ────────────────────────────────────────────────────────
        green_fill  = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        bold_font   = Font(bold=True)
        wrap_centre = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # ── Summary rows 3–4 (Cr./Dr. counts and totals) ─────────────────
        last_row = 6 + len(self)
        ws['M3'] = 'Cr.'
        ws['N3'] = f'=COUNTIF(H7:H{last_row},0)'
        ws['O3'] = f'=SUMIF(H7:H{last_row},0,J7:J{last_row})/100'
        ws['M4'] = 'Dr.'
        ws['N4'] = f'=COUNTIF(H7:H{last_row},1)'
        ws['O4'] = f'=SUMIF(H7:H{last_row},1,J7:J{last_row})/100'
        ws['L5'] = '=IF(O3=O4,"","Debits & Credits Differs, please check before converting")'

        # ── Header row 6 ──────────────────────────────────────────────────
        headers = [
            ('A', 'Tran ID (04)',                                      False),
            ('B', 'Destination Bank (04)',                             False),
            ('C', 'Destination Br   (03)',                             False),
            ('D', 'Destination  Account                (12)',          False),
            ('E', 'Destination Account Name (20)',                     False),
            ('F', 'TRN Code (02)',                                     False),
            ('G', 'Return Code (02)',                                  False),
            ('H', 'Cr/ Dr Code (01)',                                  False),
            ('I', 'Return Date (06)',                                  False),
            ('J', 'Amount (12)',                                       False),
            ('K', 'Currency Code (03)',                                True),
            ('L', 'Originating Bank (04)',                             True),
            ('M', 'Originating Barnch (03)',                           True),
            ('N', 'Originating Account                            (12)', True),
            ('O', 'Originating Account    Name (20)',                  True),
            ('P', 'Purticulars                         (15)',          True),
            ('Q', 'Reference                            (15)',         True),
            ('R', 'Value Date     (YYMMDD) (06)',                      True),
            ('S', 'Security Field           (06)',                     True),
            ('T', 'Filler (01)',                                       True),
        ]
        ws.row_dimensions[6].height = 81
        for col_letter, header_text, green in headers:
            cell = ws[f'{col_letter}6']
            cell.value = header_text
            cell.font  = bold_font
            cell.alignment = wrap_centre
            if green:
                cell.fill = green_fill

        # ── Column widths (matching original) ────────────────────────────
        col_widths = {
            'A': 4.66, 'C': 3.66, 'D': 12.66, 'E': 20.66, 'F': 2.66,
            'H': 1.66, 'I': 6.66, 'J': 12.66, 'K': 3.66,  'L': 4.66,
            'M': 3.66, 'N': 12.66,'O': 20.66, 'P': 15.66, 'R': 6.66,
            'T': 1.66,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # ── Data rows starting at row 7 ───────────────────────────────────
        def digits(val):
            return re.sub(r'\D', '', str(val or '')) or '0'

        def safe_int(val):
            d = digits(val)
            return int(d) if d else 0

        for i, slip in enumerate(self):
            r = 7 + i
            bank = slip.employee_id.bank_account_id

            dest_bank   = safe_int(slip.x_dest_bank_micr)
            dest_branch = safe_int(slip.x_dest_branch)
            dest_acc    = safe_int(bank.acc_number if bank else '')
            dest_name   = (slip.employee_id.name or '')[:20]
            trn         = safe_int(slip.x_trn_code) or 23
            ret_code    = safe_int(slip.x_return_code)
            cr_dr       = safe_int(slip.x_cr_dr_code)
            ret_date    = safe_int(slip.x_return_date)
            amount_cents = int(round((slip.net_wage or 0) * 100))
            currency    = (slip.x_currency_code or 'SLR')[:3]
            orig_bank   = safe_int(slip.x_orig_bank_micr)
            orig_branch = safe_int(slip.x_orig_branch)
            orig_acc    = safe_int(slip.x_orig_account)
            orig_name   = (slip.x_orig_name or '')[:20]
            particulars = slip.x_particulars or ''
            reference   = slip.x_cbc_reference or ''
            val_date    = slip.x_value_date or ''

            row_data = {
                'A': (0,           '0000'),
                'B': (dest_bank,   '0000'),
                'C': (dest_branch, '000'),
                'D': (dest_acc,    '000000000000'),
                'E': (dest_name,   'General'),
                'F': (trn,         'General'),
                'G': (ret_code,    '00'),
                'H': (cr_dr,       'General'),
                'I': (ret_date,    '000000'),
                'J': (amount_cents,'000000000000'),
                'K': (currency,    'General'),
                'L': (orig_bank,   '0000'),
                'M': (orig_branch, '000'),
                'N': (orig_acc,    '000000000000'),
                'O': (orig_name,   'General'),
                'P': (particulars, 'General'),
                'Q': (reference,   'General'),
                'R': (val_date,    'General'),
                'S': ('',          'General'),
                'T': ('@',         'General'),
            }
            for col_letter, (value, fmt) in row_data.items():
                cell = ws[f'{col_letter}{r}']
                cell.value = value
                if fmt != 'General':
                    cell.number_format = fmt

        # ── Serialize ─────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)

        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'PAYMASTER_BANK_DATA.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(buf.getvalue()).decode(),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': 'hr.payslip',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
